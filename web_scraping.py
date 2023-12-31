import tkinter as tk
from tkinter import messagebox
import os
import shutil
import pandas as pd
from time import sleep
import requests
import locale
from datetime import datetime
import yfinance as yf
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Variável global para definição de usuário
user = os.getlogin()

# Variável global para rastrear a janela atual
current_window = None

# Defina a lista de empresas desejadas
empresas = ['BBDC4', 'ITUB4', 'PETR4']


def avancar_window(window):
    window.destroy()
    messagebox.showinfo('Mensagem', 'Arquivo Criado com Sucesso!')

def processamento_empresa(empresa, caminho_arquivo):
    empresa_str = str(empresa) + '.SA'
    print(empresa_str)
    cotacao = yf.download(empresa_str, start='2022-01-01', end='2023-01-01')

    try:
        # Carrega o livro existente
        book = load_workbook(caminho_arquivo)
    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo livro
        book = Workbook()

    # Cria um escritor Pandas ExcelWriter usando o livro existente
    writer = pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace')

    # Adiciona as empresas à planilha 'Empresas' se não existir
    if 'Empresas' not in book.sheetnames:
        df_empresas = pd.DataFrame({'Empresas': empresas})
        df_empresas.to_excel(writer, sheet_name='Empresas', index=False)

    # Verifica se a planilha da empresa já existe no arquivo
    if empresa in book.sheetnames:
        del book[empresa]

    cotacao.to_excel(writer, sheet_name=empresa, index=True)

    cotacao['Adj Close'].plot(figsize=(15, 10))
    plt.title(empresa)
    img_path = f'{empresa}.png'
    plt.savefig(img_path)
    plt.close()

    sheet = writer.sheets[empresa]
    img = Image(img_path)
    sheet.add_image(img, f'{get_column_letter(cotacao.shape[1] + 3)}1')

    # Salva as alterações no arquivo Excel
    writer.save()


def handle_advance_button_analise(window, checkbox1_var, checkbox2_var):
    if checkbox1_var.get() or checkbox2_var.get():
        caminho_arquivo = rf'C:\Users\{user}\Downloads\Empresas.xlsx'

        for empresa in empresas:
            processamento_empresa(empresa, caminho_arquivo)

        if checkbox1_var.get():
            os.startfile(caminho_arquivo)

    messagebox.showinfo('Sucesso', 'Análise de mercado realizada com sucesso!')
    close_current_window()    
    
def open_analise_window():
    global current_window
    if current_window:
        current_window.destroy()
    current_window = tk.Toplevel(janela)
    current_window.title('Análise de Mercado')
    current_window.geometry('300x250')

    orientacao_label = tk.Label(current_window, text='Selecione as opções desejadas:')
    orientacao_label.pack(pady=10)

    checkbox1_var = tk.BooleanVar(value=False)
    checkbox2_var = tk.BooleanVar(value=False)

    def on_checkbox1_change():
        if checkbox1_var.get():
            checkbox2_var.set(False)

    def on_checkbox2_change():
        if checkbox2_var.get():
            checkbox1_var.set(False)

    checkbox1 = tk.Checkbutton(current_window, text='Realizar download e visualizar arquivo', variable=checkbox1_var, command=on_checkbox1_change)
    checkbox1.pack()

    checkbox2 = tk.Checkbutton(current_window, text='Apenas realizar download localmente', variable=checkbox2_var, command=on_checkbox2_change)
    checkbox2.pack()

    button_frame = tk.Frame(current_window)
    button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    retroceder_button = tk.Button(button_frame, text='Retroceder para Principal', command=close_current_window)
    retroceder_button.pack(side=tk.LEFT)

    avancar_button = tk.Button(button_frame, text='Avançar', command=lambda: handle_advance_button_analise(current_window, checkbox1_var, checkbox2_var))
    avancar_button.pack(side=tk.RIGHT)

    # Verifica se o diretório existe, se não existir, cria
    download_dir = rf'C:\Users\{user}\Downloads'
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)

    # Verifica se o arquivo 'Empresas.xlsx' existe, se não existir, cria um arquivo vazio
    empresas_xlsx_path = os.path.join(download_dir, 'Empresas.xlsx')
    if not os.path.exists(empresas_xlsx_path):
        with pd.ExcelWriter(empresas_xlsx_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=['Empresas']).to_excel(writer, sheet_name='Empresas', index=False)

def log_cot():
    # Definindo o local para o formato de moeda
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    # Verifica se o arquivo existe
    if not os.path.isfile(r'C:\Users\{}\Downloads\Cotações.xlsx'.format(user)):
        # Cria um DataFrame vazio
        data = {'Moeda': ['Dólar', 'Euro', 'Bitcoin'], 'Cotação': [0, 0, 0], 'Data Última Atualização': ['N/A', 'N/A', 'N/A']}
        df = pd.DataFrame(data)

        # Salva o DataFrame em um arquivo Excel
        df.to_excel(r'C:\Users\{}\Downloads\Cotações.xlsx'.format(user), index=False)

    try:
        requisicao = requests.get('https://economia.awesomeapi.com.br/last/USD-BRL,EUR-BRL,BTC-BRL')
        requisicao_dic = requisicao.json()

        cotacao_dolar = locale.currency(float(requisicao_dic['USDBRL']['bid']), grouping=True, symbol=None)
        cotacao_euro = locale.currency(float(requisicao_dic['EURBRL']['bid']), grouping=True, symbol=None)
        cotacao_btc = locale.currency(float(requisicao_dic['BTCBRL']['bid']) * 1000, grouping=True, symbol=None)

        # Leitura da tabela
        tabela = pd.read_excel(r'C:\Users\{}\Downloads\Cotações.xlsx'.format(user))

        # Atualização das cotações
        tabela.loc[0, 'Cotação'] = cotacao_dolar
        tabela.loc[1, 'Cotação'] = cotacao_euro
        tabela.loc[2, 'Cotação'] = cotacao_btc

        # Atualização da data e hora
        tabela.loc[0, 'Data Última Atualização'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Escrita da tabela
        tabela.to_excel(r'C:\Users\{}\Downloads\Cotações.xlsx'.format(user), index=False)
        print(f'Cotação Atualizada. {datetime.now()}')

    except requests.exceptions.RequestException as e:
        print(f"Erro na requisição: {e}")
    except pd.errors.EmptyDataError as e:
        print(f"Erro na leitura da planilha: {e}")
    except pd.errors.ParserError as e:
        print(f"Erro na leitura da planilha: {e}")

def open_cot_window():
    global current_window
    if current_window:
        current_window.destroy()
    current_window = tk.Toplevel(janela)
    current_window.title('Pesquisar Cotações de Moeda')
    current_window.geometry('300x250')

    orientacao_label = tk.Label(current_window, text='Selecione as opções desejadas:')
    orientacao_label.pack(pady=10)

    checkbox1_var = tk.BooleanVar(value=False)
    checkbox2_var = tk.BooleanVar(value=False)

    def on_checkbox1_change():
        if checkbox1_var.get():
            checkbox2_var.set(False)

    def on_checkbox2_change():
        if checkbox2_var.get():
            checkbox1_var.set(False)

    checkbox1 = tk.Checkbutton(current_window, text='Realizar download e vizualizar arquivo', variable=checkbox1_var, command=on_checkbox1_change)
    checkbox1.pack()

    checkbox2 = tk.Checkbutton(current_window, text='Apenas realizar download localmente', variable=checkbox2_var, command=on_checkbox2_change)
    checkbox2.pack()

    button_frame = tk.Frame(current_window)
    button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

    retroceder_button = tk.Button(button_frame, text='Retroceder para Principal', command=close_current_window)
    retroceder_button.pack(side=tk.LEFT)

    avancar_button = tk.Button(button_frame, text='Avançar', command=lambda: handle_advance_button(current_window, checkbox1_var, checkbox2_var))
    avancar_button.pack(side=tk.RIGHT)

def handle_advance_button(window, checkbox1_var, checkbox2_var):
    # Verificar o estado das caixas de seleção
    if checkbox1_var.get():
        caminho_arquivo = r'C:\Users\{}\Downloads\Cotações.xlsx'.format(user)
        log_cot()
        os.startfile(caminho_arquivo)

    if checkbox2_var.get():
        log_cot()

    # Mostrando uma mensagem de sucesso e fechando a janela secundária
    messagebox.showinfo('Sucesso', 'Cotação atualizada com sucesso!')
    close_current_window()

def close_current_window():
    global current_window
    if current_window:
        current_window.destroy()
    current_window = None

# Criando Janela Principal
janela = tk.Tk()
janela.title('Web Scraping')
janela.geometry('360x400')

# Centralizando orientação de texto usando pack
texto_orientacao = tk.Label(janela, text='Escolha qual dado deseja buscar na Web:')
texto_orientacao.pack(pady=10)  # Adicionamos pady para espaçamento vertical

# Centralizando os botões usando pack
botao = tk.Button(janela, text='Análise de Mercado', command=open_analise_window)
botao.pack(pady=5)  # Adicionamos pady para espaçamento vertical

botao_3 = tk.Button(janela, text='Pesquisar Cotações de Moeda', command=open_cot_window)
botao_3.pack(pady=5)  # Adicionamos pady para espaçamento vertical

janela.mainloop()

#By Matheus Siqueira
